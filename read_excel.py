import openpyxl
import os
import numpy as np

def load_excel_data(path:str|os.PathLike,sheet:str):
    wb = openpyxl.load_workbook(path,data_only=True)
    s = wb[sheet]
    row_no = s.max_row
    col_no = s.max_column
    data = (np.zeros((row_no,col_no)))
    row_is_nan_no = 0
    delete_nan_row = np.zeros((1,row_no))
    for row in s.iter_rows(min_row=1, max_col=col_no, max_row=row_no):
        for cell in row:
            if isinstance(cell.value,float) or isinstance(cell.value,int):
                data[cell.row-1][cell.column-1] = float(cell.value)
                row_is_nan_no = 0
            else:
                data[cell.row-1][cell.column-1] = np.nan
                row_is_nan_no+=1
        if row_is_nan_no==row_no:
            delete_nan_row[row[0].row-1]=1
    data=np.delete(data,np.argmin(abs(delete_nan_row-1),),axis=0)
    return data


if __name__=='__main__':
    import os
    # os.chdir('/content/drive/MyDrive/Colab Notebooks')  # Colab 換路徑使用

    import openpyxl
    wb = openpyxl.load_workbook('excel-read.xlsx',data_only=True)     # 開啟 Excel 檔案

    names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
    s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
    s2 = wb['工作表2']         # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

    # print(names)
    # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
    # print(s1.title, s1.max_row, s1.max_column)
    # print(s2.title, s2.max_row, s2.max_column)
    # print(s1['A1'].value)        # 取出 A1 的內容
    # print(s1.cell(1, 1).value)   # 等同取出 A1 的內容
    # print(s2['B2'].value)        # 取出 B2 的內容
    # print(s2.cell(2, 2).value)   # 等同取出 B2 的內容
    for col in s1.iter_cols(min_row=1, max_col=s1.max_column, max_row=s1.max_row):
        for cell in col:
            print(cell.value)
        # print(col.value)
    D=load_excel_data("D:/Homework/Master_degree/ring/CMT/nonlinear/ring spectrum.xlsx", '0dbm')
    print(D)
    # print(D[:,2])
